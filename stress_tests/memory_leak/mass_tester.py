#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ЭКСТРЕМАЛЬНЫЙ параллельный массовый тестер для нагрузочного тестирования API
10000 итераций ОДНОВРЕМЕННО: каждая = 3 WebAccount + 4 DMZ запроса параллельно
ИТОГО: 70,000 одновременных запросов!
ТОЧЕЧНАЯ АТАКА - все запросы стартуют в один момент!
"""

import time
import threading
import json
from datetime import datetime
import os

from webaccount_client import WebAccountClient
from dmz_client import DmzClient

class PointAttackTester:
    def __init__(self, concurrent_iterations=10000):
        self.concurrent_iterations = concurrent_iterations
        self.total_requests = concurrent_iterations * 7  # 3 WA + 4 DMZ
        
        # Статистика
        self.stats = {
            'webaccount': {'success': 0, 'error': 0, 'total_time': 0, 'data_received': 0},
            'dmz': {'success': 0, 'error': 0, 'total_time': 0, 'data_received': 0},
            'completed_iterations': 0,
            'total_requests_sent': 0,
            'attack_start_time': None,
            'first_response_time': None,
            'last_response_time': None
        }
        
        # Детальные данные каждой итерации
        self.detailed_data = []
        
        # Блокировки для безопасного обновления
        self.stats_lock = threading.Lock()
        self.data_lock = threading.Lock()
        
        # Время начала тестирования
        self.test_start_time = None
        
        # ОДИН ГЛОБАЛЬНЫЙ БАРЬЕР ДЛЯ ВСЕХ ЗАПРОСОВ!
        self.global_barrier = threading.Barrier(self.total_requests)
        self.attack_moment = None
    
    def _calculate_internet_speed(self, data_size, time_taken):
        """Вычисляет скорость интернета в Mbps"""
        if time_taken <= 0:
            return 0
        # data_size в байтах, time_taken в секундах
        # Скорость = (байты * 8) / (секунды * 1000000) = Mbps
        speed_mbps = (data_size * 8) / (time_taken * 1000000)
        return round(speed_mbps, 3)
    
    def _single_webaccount_request(self, endpoint, iteration):
        """Выполняет один запрос к WebAccount endpoint с детальными метриками"""
        client = WebAccountClient()
        
        try:
            # ЖДЕМ ГЛОБАЛЬНОЙ КОМАНДЫ ДЛЯ ОДНОВРЕМЕННОГО СТАРТА ВСЕХ 70,000 ЗАПРОСОВ!
            self.global_barrier.wait()
            
            # ВСЕ 70,000 ЗАПРОСОВ СТАРТУЮТ ОДНОВРЕМЕННО ЗДЕСЬ!
            if self.attack_moment is None:
                with self.stats_lock:
                    if self.attack_moment is None:
                        self.attack_moment = time.time()
            
            # Время начала запроса (после барьера)
            request_start = time.time()
            
            result = client.call_endpoint(endpoint)
            
            # Время окончания запроса
            request_end = time.time()
            
            # Обновляем статистику времени первого/последнего ответа
            with self.stats_lock:
                if self.stats['first_response_time'] is None:
                    self.stats['first_response_time'] = request_end
                self.stats['last_response_time'] = request_end
            
            # Вычисляем метрики
            request_time = request_end - request_start
            data_size = len(result['data'] or '')
            internet_speed = self._calculate_internet_speed(data_size, request_time)
            
            return {
                'type': 'webaccount',
                'endpoint': endpoint,
                'iteration': iteration,
                'success': result['success'],
                'error': result['error'],
                'request_start_time': request_start,
                'request_end_time': request_end,
                'request_time': round(request_time, 4),
                'response_time': round(request_time, 4),
                'data_size': data_size,
                'internet_speed_mbps': internet_speed,
                'attack_delay': round(request_start - self.attack_moment, 4) if self.attack_moment else 0,
                'timestamp': datetime.now().isoformat()
            }
        except Exception as e:
            # Даже при ошибке ждем барьер
            try:
                self.global_barrier.wait()
            except:
                pass
                
            if self.attack_moment is None:
                with self.stats_lock:
                    if self.attack_moment is None:
                        self.attack_moment = time.time()
            
            request_end = time.time()
            request_start = request_end  # При ошибке время минимальное
            request_time = 0.001
            
            return {
                'type': 'webaccount',
                'endpoint': endpoint,
                'iteration': iteration,
                'success': False,
                'error': str(e),
                'request_start_time': request_start,
                'request_end_time': request_end,
                'request_time': round(request_time, 4),
                'response_time': round(request_time, 4),
                'data_size': 0,
                'internet_speed_mbps': 0,
                'attack_delay': round(request_start - self.attack_moment, 4) if self.attack_moment else 0,
                'timestamp': datetime.now().isoformat()
            }
        finally:
            client.close()
    
    def _single_dmz_request(self, endpoint, iteration):
        """Выполняет один запрос к DMZ endpoint с детальными метриками"""
        client = DmzClient()
        
        try:
            # ЖДЕМ ГЛОБАЛЬНОЙ КОМАНДЫ ДЛЯ ОДНОВРЕМЕННОГО СТАРТА ВСЕХ 70,000 ЗАПРОСОВ!
            self.global_barrier.wait()
            
            # ВСЕ 70,000 ЗАПРОСОВ СТАРТУЮТ ОДНОВРЕМЕННО ЗДЕСЬ!
            if self.attack_moment is None:
                with self.stats_lock:
                    if self.attack_moment is None:
                        self.attack_moment = time.time()
            
            # Время начала запроса (после барьера)
            request_start = time.time()
            
            result = client.call_endpoint(endpoint)
            
            # Время окончания запроса
            request_end = time.time()
            
            # Обновляем статистику времени первого/последнего ответа
            with self.stats_lock:
                if self.stats['first_response_time'] is None:
                    self.stats['first_response_time'] = request_end
                self.stats['last_response_time'] = request_end
            
            # Вычисляем метрики
            request_time = request_end - request_start
            data_size = len(result['data'] or '')
            internet_speed = self._calculate_internet_speed(data_size, request_time)
            
            return {
                'type': 'dmz',
                'endpoint': endpoint,
                'iteration': iteration,
                'success': result['success'],
                'error': result['error'],
                'request_start_time': request_start,
                'request_end_time': request_end,
                'request_time': round(request_time, 4),
                'response_time': round(request_time, 4),
                'data_size': data_size,
                'internet_speed_mbps': internet_speed,
                'attack_delay': round(request_start - self.attack_moment, 4) if self.attack_moment else 0,
                'timestamp': datetime.now().isoformat()
            }
        except Exception as e:
            # Даже при ошибке ждем барьер
            try:
                self.global_barrier.wait()
            except:
                pass
                
            if self.attack_moment is None:
                with self.stats_lock:
                    if self.attack_moment is None:
                        self.attack_moment = time.time()
            
            request_end = time.time()
            request_start = request_end
            request_time = 0.001
            
            return {
                'type': 'dmz',
                'endpoint': endpoint,
                'iteration': iteration,
                'success': False,
                'error': str(e),
                'request_start_time': request_start,
                'request_end_time': request_end,
                'request_time': round(request_time, 4),
                'response_time': round(request_time, 4),
                'data_size': 0,
                'internet_speed_mbps': 0,
                'attack_delay': round(request_start - self.attack_moment, 4) if self.attack_moment else 0,
                'timestamp': datetime.now().isoformat()
            }
        finally:
            client.close()
    
    def _update_stats(self, result):
        """Безопасно обновляет статистику"""
        with self.stats_lock:
            api_type = result['type']
            
            if result['error'] and not result['success']:
                self.stats[api_type]['error'] += 1
            elif result['success']:
                self.stats[api_type]['success'] += 1
            
            self.stats[api_type]['total_time'] += result['request_time']
            self.stats[api_type]['data_received'] += result['data_size']
    
    def _save_detailed_data(self, result):
        """Безопасно сохраняет один результат"""
        with self.data_lock:
            # Добавляем относительное время с начала теста
            if self.test_start_time:
                result['time_from_start'] = round(result['request_start_time'] - self.test_start_time, 4)
            
            self.detailed_data.append(result)
    
    def _run_single_request(self, request_type, endpoint, iteration):
        """Выполняет один запрос с автоматическим сохранением"""
        if request_type == 'webaccount':
            result = self._single_webaccount_request(endpoint, iteration)
        else:  # dmz
            result = self._single_dmz_request(endpoint, iteration)
        
        # Обновляем статистику и сохраняем данные
        self._update_stats(result)
        self._save_detailed_data(result)
        
        return result
    
    def _save_to_files(self):
        """Сохраняет данные в различные форматы файлов"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # 1. JSON файл с полными данными
        json_filename = f'stress_test_detailed_{timestamp}.json'
        with open(json_filename, 'w', encoding='utf-8') as f:
            json.dump({
                'test_info': {
                    'concurrent_iterations': self.concurrent_iterations,
                    'total_requests': len(self.detailed_data),
                    'test_start_time': self.test_start_time,
                    'attack_moment': self.attack_moment,
                    'first_response_time': self.stats.get('first_response_time'),
                    'last_response_time': self.stats.get('last_response_time'),
                    'timestamp': timestamp
                },
                'summary_stats': self.stats,
                'detailed_data': self.detailed_data
            }, f, indent=2, ensure_ascii=False)
        
        # 2. Excel файл для удобного анализа
        excel_filename = f'stress_test_summary_{timestamp}.xlsx'
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Point Attack Results"
            
            # Заголовки
            headers = [
                'iteration', 'type', 'endpoint', 'success', 'error',
                'request_time', 'response_time', 'data_size', 'internet_speed_mbps',
                'attack_delay', 'time_from_start', 'timestamp'
            ]
            
            # Записываем заголовки с форматированием
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Записываем данные
            for row, data in enumerate(self.detailed_data, 2):
                for col, header in enumerate(headers, 1):
                    value = data.get(header, '')
                    ws.cell(row=row, column=col, value=value)
            
            # Автоширина столбцов
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(excel_filename)
            
        except ImportError:
            # Fallback к CSV если openpyxl не установлен
            excel_filename = f'stress_test_summary_{timestamp}.csv'
            import csv
            with open(excel_filename, 'w', newline='', encoding='utf-8') as f:
                if self.detailed_data:
                    fieldnames = [
                        'iteration', 'type', 'endpoint', 'success', 'error',
                        'request_time', 'response_time', 'data_size', 'internet_speed_mbps',
                        'attack_delay', 'time_from_start', 'timestamp'
                    ]
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writeheader()
                    
                    for row in self.detailed_data:
                        csv_row = {k: v for k, v in row.items() if k in fieldnames}
                        writer.writerow(csv_row)
        
        # 3. Сводный отчет в текстовом формате
        report_filename = f'stress_test_report_{timestamp}.txt'
        with open(report_filename, 'w', encoding='utf-8') as f:
            f.write("=" * 70 + "\n")
            f.write("ОТЧЕТ ТОЧЕЧНОЙ АТАКИ - ЭКСТРЕМАЛЬНОЕ НАГРУЗОЧНОЕ ТЕСТИРОВАНИЕ\n")
            f.write("=" * 70 + "\n\n")
            
            f.write(f"Дата и время: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Одновременных итераций: {self.concurrent_iterations:,}\n")
            f.write(f"Общее количество запросов: {len(self.detailed_data):,}\n\n")
            
            # Статистика атаки
            if self.attack_moment and self.stats.get('first_response_time') and self.stats.get('last_response_time'):
                first_delay = self.stats['first_response_time'] - self.attack_moment
                last_delay = self.stats['last_response_time'] - self.attack_moment
                attack_duration = self.stats['last_response_time'] - self.stats['first_response_time']
                
                f.write("СТАТИСТИКА ТОЧЕЧНОЙ АТАКИ:\n")
                f.write("-" * 30 + "\n")
                f.write(f"Время до первого ответа: {first_delay:.4f}с\n")
                f.write(f"Время до последнего ответа: {last_delay:.4f}с\n") 
                f.write(f"Продолжительность атаки: {attack_duration:.4f}с\n\n")
            
            # Статистика по API
            f.write("СТАТИСТИКА ПО API:\n")
            f.write("-" * 30 + "\n")
            for api_type, stats in self.stats.items():
                if api_type in ['completed_iterations', 'total_requests_sent', 'attack_start_time', 'first_response_time', 'last_response_time']:
                    continue
                f.write(f"{api_type.upper()}:\n")
                f.write(f"  Успешных: {stats['success']:,}\n")
                f.write(f"  Ошибок: {stats['error']:,}\n")
                f.write(f"  Данных получено: {stats['data_received']:,} байт\n")
                f.write(f"  Общее время: {stats['total_time']:.2f}с\n\n")
            
            # Средние показатели
            if self.detailed_data:
                successful_requests = [r for r in self.detailed_data if r['success']]
                if successful_requests:
                    avg_request_time = sum(r['request_time'] for r in successful_requests) / len(successful_requests)
                    avg_speed = sum(r['internet_speed_mbps'] for r in successful_requests) / len(successful_requests)
                    avg_data_size = sum(r['data_size'] for r in successful_requests) / len(successful_requests)
                    avg_attack_delay = sum(r.get('attack_delay', 0) for r in successful_requests) / len(successful_requests)
                    
                    f.write("СРЕДНИЕ ПОКАЗАТЕЛИ (успешные запросы):\n")
                    f.write("-" * 30 + "\n")
                    f.write(f"Среднее время запроса: {avg_request_time:.4f}с\n")
                    f.write(f"Средняя скорость: {avg_speed:.3f} Mbps\n")
                    f.write(f"Средний размер данных: {avg_data_size:.0f} байт\n")
                    f.write(f"Средняя задержка атаки: {avg_attack_delay:.4f}с\n\n")
            
            f.write("ФАЙЛЫ С ДАННЫМИ:\n")
            f.write("-" * 30 + "\n")
            f.write(f"Детальные данные (JSON): {json_filename}\n")
            f.write(f"Сводная таблица (Excel): {excel_filename}\n")
            f.write(f"Этот отчет: {report_filename}\n")
        
        return json_filename, excel_filename, report_filename
    
    def run_point_attack(self):
        """Запускает ТОЧЕЧНУЮ АТАКУ - все запросы стартуют ОДНОВРЕМЕННО!"""
        print("💥 ЗАПУСК ТОЧЕЧНОЙ АТАКИ - ОДНОВРЕМЕННЫЙ СТАРТ ВСЕХ ЗАПРОСОВ!")
        print("=" * 70)
        print(f"📅 Время начала: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"🎯 Одновременных итераций: {self.concurrent_iterations:,}")
        print(f"🔄 В каждой итерации: 3 WebAccount + 4 DMZ запроса")
        print(f"💥 ОБЩЕЕ КОЛИЧЕСТВО ОДНОВРЕМЕННЫХ ЗАПРОСОВ: {self.total_requests:,}")
        print("⚡ ВСЕ ЗАПРОСЫ СТАРТУЮТ В ОДИН МОМЕНТ!")
        print("=" * 70)
        
        self.test_start_time = time.time()
        
        print(f"🚀 Подготовка {self.total_requests:,} запросов...")
        print("💫 Создание потоков для точечной атаки...")
        
        # Получаем списки endpoints
        wa_client = WebAccountClient()
        dmz_client = DmzClient()
        wa_endpoints = wa_client.endpoints
        dmz_endpoints = dmz_client.endpoints
        wa_client.close()
        dmz_client.close()
        
        # Создаем ВСЕ потоки для ВСЕХ запросов СРАЗУ
        threads = []
        
        for iteration in range(1, self.concurrent_iterations + 1):
            # WebAccount endpoints для этой итерации
            for endpoint in wa_endpoints:
                thread = threading.Thread(
                    target=self._run_single_request, 
                    args=('webaccount', endpoint, iteration)
                )
                threads.append(thread)
            
            # DMZ endpoints для этой итерации  
            for endpoint in dmz_endpoints:
                thread = threading.Thread(
                    target=self._run_single_request, 
                    args=('dmz', endpoint, iteration)
                )
                threads.append(thread)
        
        print("⚡ ЗАПУСК ВСЕХ ПОТОКОВ...")
        start_time = time.time()
        
        # Запускаем ВСЕ потоки одновременно - они будут ждать на барьере
        for thread in threads:
            thread.start()
        
        print(f"💥 {len(threads):,} ПОТОКОВ ЗАПУЩЕНО!")
        print("⏳ ВСЕ ПОТОКИ ЖДУТ КОМАНДЫ НА ОДНОВРЕМЕННЫЙ СТАРТ...")
        print("📊 Ожидаю завершения всех запросов...")
        
        # Отслеживаем прогресс по количеству записанных результатов
        completed = 0
        while completed < self.total_requests:
            time.sleep(0.5)  # Проверяем каждые 0.5 секунд
            with self.data_lock:
                new_completed = len(self.detailed_data)
            
            if new_completed > completed:
                completed = new_completed
                elapsed = time.time() - start_time
                rate = completed / elapsed if elapsed > 0 else 0
                percentage = (completed / self.total_requests) * 100
                
                if completed % 5000 == 0 or completed <= 50:
                    print(f"📈 Завершено: {completed:,}/{self.total_requests:,} ({percentage:.1f}%) - {rate:.1f} req/sec")
        
        # Ждем завершения всех потоков
        for thread in threads:
            thread.join()
        
        total_time = time.time() - start_time
        
        # Статистика атаки
        if self.attack_moment:
            attack_preparation_time = self.attack_moment - start_time
            print(f"\n💥 ТОЧЕЧНАЯ АТАКА ЗАВЕРШЕНА!")
            print(f"⏱️ Время подготовки: {attack_preparation_time:.4f}с")
            print(f"⏱️ Общее время: {total_time:.2f}с")
            
            if self.stats.get('first_response_time') and self.stats.get('last_response_time'):
                first_response_delay = self.stats['first_response_time'] - self.attack_moment
                last_response_delay = self.stats['last_response_time'] - self.attack_moment
                attack_duration = self.stats['last_response_time'] - self.stats['first_response_time']
                print(f"⚡ Первый ответ через: {first_response_delay:.4f}с после атаки")
                print(f"🏁 Последний ответ через: {last_response_delay:.4f}с после атаки")
                print(f"🕐 Продолжительность атаки: {attack_duration:.4f}с")
        
        # Сохраняем данные в файлы
        print("\n💾 Сохраняю детальные данные в файлы...")
        json_file, excel_file, report_file = self._save_to_files()
        print(f"✅ Данные сохранены:")
        print(f"   📋 JSON: {json_file}")
        print(f"   📊 Excel: {excel_file}")
        print(f"   📄 Отчет: {report_file}")
        
        return total_time
    
    def print_final_stats(self, total_time):
        """Выводит финальную статистику"""
        print("\n" + "="*70)
        print("📈 ФИНАЛЬНАЯ СТАТИСТИКА ТОЧЕЧНОЙ АТАКИ")
        print("="*70)
        
        # WebAccount статистика
        wa_stats = self.stats['webaccount']
        wa_expected = self.concurrent_iterations * 3
        print(f"\n🔐 WEBACCOUNT API:")
        print(f"   📊 Ожидалось запросов: {wa_expected:,}")
        print(f"   ✅ Успешных: {wa_stats['success']:,}")
        print(f"   ❌ Ошибок: {wa_stats['error']:,}")
        print(f"   📦 Данных получено: {wa_stats['data_received']:,} символов")
        if wa_expected > 0:
            print(f"   📊 Среднее время ответа: {wa_stats['total_time'] / wa_expected:.3f}с")
        
        # DMZ статистика
        dmz_stats = self.stats['dmz']
        dmz_expected = self.concurrent_iterations * 4
        print(f"\n🌐 DMZ API:")
        print(f"   📊 Ожидалось запросов: {dmz_expected:,}")
        print(f"   ✅ Успешных: {dmz_stats['success']:,}")
        print(f"   ❌ Ошибок: {dmz_stats['error']:,}")
        print(f"   📦 Данных получено: {dmz_stats['data_received']:,} символов")
        if dmz_expected > 0:
            print(f"   📊 Среднее время ответа: {dmz_stats['total_time'] / dmz_expected:.3f}с")
        
        # Общая статистика
        total_expected = wa_expected + dmz_expected
        total_success = wa_stats['success'] + dmz_stats['success']
        total_errors = wa_stats['error'] + dmz_stats['error']
        total_data = wa_stats['data_received'] + dmz_stats['data_received']
        
        print(f"\n🎯 ОБЩАЯ СТАТИСТИКА:")
        print(f"   📊 Ожидалось запросов: {total_expected:,}")
        print(f"   ✅ Успешных запросов: {total_success:,}")
        print(f"   ❌ Ошибок: {total_errors:,}")
        print(f"   ⏱️ Общее время: {total_time:.2f}с")
        
        if total_time > 0:
            print(f"   🚀 Скорость запросов: {total_success / total_time:.1f} req/sec")
        
        print(f"   📦 Всего данных: {total_data:,} символов ({total_data / 1024 / 1024:.1f} MB)")
        
        # Эффективность
        if total_expected > 0:
            success_rate = (total_success / total_expected) * 100
            print(f"   🎯 Успешность запросов: {success_rate:.1f}%")
        
        # Детальная статистика атаки
        if self.detailed_data:
            successful_requests = [r for r in self.detailed_data if r['success']]
            if successful_requests:
                attack_delays = [r.get('attack_delay', 0) for r in successful_requests]
                avg_attack_delay = sum(attack_delays) / len(attack_delays)
                max_attack_delay = max(attack_delays)
                min_attack_delay = min(attack_delays)
                
                print(f"\n⚡ СИНХРОНИЗАЦИЯ АТАКИ:")
                print(f"   📊 Средняя задержка старта: {avg_attack_delay:.4f}с")
                print(f"   ⬆️ Максимальная задержка: {max_attack_delay:.4f}с")
                print(f"   ⬇️ Минимальная задержка: {min_attack_delay:.4f}с")
                
                # Статистика скоростей
                speeds = [r.get('internet_speed_mbps', 0) for r in successful_requests if r.get('internet_speed_mbps', 0) > 0]
                if speeds:
                    avg_speed = sum(speeds) / len(speeds)
                    max_speed = max(speeds)
                    min_speed = min(speeds)
                    
                    print(f"\n🌐 СКОРОСТЬ ИНТЕРНЕТА:")
                    print(f"   📊 Средняя: {avg_speed:.3f} Mbps")
                    print(f"   ⬆️ Максимальная: {max_speed:.3f} Mbps")
                    print(f"   ⬇️ Минимальная: {min_speed:.3f} Mbps")
        
        print("\n" + "="*70)
    
    def run_full_test(self):
        """Запускает полную точечную атаку"""
        total_time = self.run_point_attack()
        self.print_final_stats(total_time)
        
        print(f"💥 ТОЧЕЧНАЯ АТАКА ЗАВЕРШЕНА за {total_time:.2f}с")
        print(f"📅 Время окончания: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

def main():
    """Главная функция"""
    print("⚙️  ТОЧЕЧНАЯ АТАКА - ЭКСТРЕМАЛЬНЫЙ МАССОВЫЙ ТЕСТЕР")
    print("💥 ВСЕ ЗАПРОСЫ СТАРТУЮТ ОДНОВРЕМЕННО!")
    print("="*60)
    
    # Параметры по умолчанию - увеличено до 10000!
    concurrent_iterations = 10000
    
    print(f"🎯 Одновременных итераций: {concurrent_iterations:,}")
    print(f"📊 В каждой итерации: 3 WebAccount + 4 DMZ запроса")
    print(f"💥 ОБЩЕЕ КОЛИЧЕСТВО ОДНОВРЕМЕННЫХ ЗАПРОСОВ: {concurrent_iterations * 7:,}")
    print(f"📋 Детальная статистика будет сохранена в Excel")
    print(f"⚡ ТОЧЕЧНАЯ АТАКА - все запросы стартуют в один момент!")
    
    # Возможность изменить параметры
    try:
        response = input("\n❓ Изменить количество одновременных итераций? (y/N): ").strip().lower()
        if response in ['y', 'yes', 'да']:
            concurrent_iterations = int(input("Одновременных итераций: ") or concurrent_iterations)
    except (ValueError, KeyboardInterrupt):
        print("Используются параметры по умолчанию")
    
    # Проверка openpyxl
    try:
        import openpyxl
        print("✅ openpyxl найден - будет создан Excel файл")
    except ImportError:
        print("⚠️ openpyxl не найден - будет создан CSV файл")
        print("💡 Для Excel установите: pip install openpyxl")
    
    # Создаем и запускаем тестер
    tester = PointAttackTester(concurrent_iterations)
    
    try:
        tester.run_full_test()
    except KeyboardInterrupt:
        print("\n⚠️ Атака прервана пользователем")
        if tester.detailed_data:
            print("💾 Сохраняю накопленные данные...")
            tester._save_to_files()
    except Exception as e:
        print(f"\n❌ Критическая ошибка: {e}")
        if tester.detailed_data:
            print("💾 Сохраняю накопленные данные...")
            tester._save_to_files()

if __name__ == "__main__":
    main() 